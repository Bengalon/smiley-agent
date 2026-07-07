# -*- coding: utf-8 -*-
"""
build.py — מחולל אתרים לעסקים מקומיים ללא אתר.
קורא data/businesses.json ומייצר:
  - sites/<slug>/index.html   אתר מלא, נגיש, RTL, מותאם אישית לכל עסק
  - sites/index.html          עמוד אינדקס (גלריית האתרים)
  - leads.xlsx                גיליון לידים מסודר עם כל הפרטים

הרצה:  python3 build.py
"""

import json
import os
import re
import html
import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "data", "businesses.json")
SITES = os.path.join(BASE, "sites")

# ----------------------------------------------------------------------------
# פלטות צבע לפי theme
# ----------------------------------------------------------------------------
PALETTES = {
    "plum":     {"primary": "#6d28d9", "dark": "#4c1d95", "accent": "#a78bfa", "ink": "#faf5ff"},
    "slate":    {"primary": "#334155", "dark": "#0f172a", "accent": "#94a3b8", "ink": "#f8fafc"},
    "rose":     {"primary": "#e11d48", "dark": "#9f1239", "accent": "#fb7185", "ink": "#fff1f2"},
    "charcoal": {"primary": "#111827", "dark": "#030712", "accent": "#f59e0b", "ink": "#fffbeb"},
    "green":    {"primary": "#15803d", "dark": "#14532d", "accent": "#4ade80", "ink": "#f0fdf4"},
    "orange":   {"primary": "#ea580c", "dark": "#9a3412", "accent": "#fb923c", "ink": "#fff7ed"},
    "amber":    {"primary": "#b45309", "dark": "#78350f", "accent": "#fbbf24", "ink": "#fffbeb"},
    "brown":    {"primary": "#8b5e34", "dark": "#5b3a1f", "accent": "#c8925e", "ink": "#faf6f1"},
}

# ----------------------------------------------------------------------------
# תוכן לפי קטגוריה
# ----------------------------------------------------------------------------
CATEGORIES = {
    "barber": {
        "hero_emoji": "💈",
        "nav": {"services": "השירותים", "gallery": "גלריה", "hours": "שעות", "contact": "צור קשר"},
        "services_title": "השירותים שלנו",
        "services_note": "התספורות והשירותים להמחשה — נשמח להתאים לך אישית.",
        "services": [
            ("✂️", "תספורת גברים", "תספורת מדויקת בהתאמה אישית, כולל גימור ועיצוב."),
            ("🧔", "עיצוב וטיפוח זקן", "קיצוב, עיצוב וטיפוח זקן עם מגבת חמה."),
            ("💇", "תספורת נשים", "תספורת, פן ועיצוב לפי הסגנון שמתאים לך."),
            ("🎨", "צבע והחלקה", "צביעה, גוונים והחלקות בחומרים איכותיים."),
            ("👦", "תספורת ילדים", "יחס סבלני ונעים גם לקטנטנים."),
            ("💧", "חפיפה ועיצוב", "חפיפה, טריטמנט ועיצוב לשיער בריא ומטופח."),
        ],
        "features": [
            ("⭐", "יחס אישי", "מכירים כל לקוח בשם ומתאימים את התספורת בדיוק לך."),
            ("🕒", "בלי להמתין", "אפשר לתאם תור מראש ולהיכנס בזמן."),
            ("🧼", "היגיינה מלאה", "כלים מעוקרים וסביבה נקייה ומסודרת."),
        ],
        "gallery": ["תספורת קלאסית", "עיצוב זקן", "פן ועיצוב", "גוונים", "לוק ערב", "אווירה במספרה"],
        "about": "היא מספרת שכונה שבה כל לקוח מקבל יחס אישי, מקצועי וחם. אנחנו מאמינים שתספורת טובה מתחילה בהקשבה — להבין מה מתאים לך, לאורח החיים ולסגנון שלך.",
    },
    "grocery": {
        "hero_emoji": "🛒",
        "nav": {"services": "המחלקות", "gallery": "אצלנו בחנות", "hours": "שעות", "contact": "צור קשר"},
        "services_title": "מה יש אצלנו",
        "services_note": "מגוון המוצרים משתנה מדי יום — תמיד טרי.",
        "services": [
            ("🥛", "מוצרי חלב טריים", "חלב, גבינות, יוגורטים וביצים במשלוח יומי."),
            ("🍞", "מאפים ולחם", "לחם טרי, לחמניות ומאפים כל בוקר."),
            ("🥬", "פירות וירקות", "תוצרת טרייה שנבחרת ביד כל יום."),
            ("🧺", "מוצרי יסוד", "כל מה שצריך למזווה — במחירים הוגנים."),
            ("🧃", "שתייה וחטיפים", "מבחר משקאות, חטיפים וכיבוד לאורחים."),
            ("🧴", "ניקיון וטואלטיקה", "מוצרי בית וטיפוח לשימוש יומיומי."),
        ],
        "features": [
            ("🚴", "משלוח לבית", "מזמינים בטלפון ומקבלים עד הדלת."),
            ("😊", "פנים מוכרות", "שירות אישי כמו של פעם, קרוב לבית."),
            ("🕗", "פתוח עד מאוחר", "כאן בשבילכם גם כשכולם כבר סגורים."),
        ],
        "gallery": ["מדף הירקות", "פינת המאפים", "מקרר המשקאות", "מוצרי חלב", "פינת הכיבוד", "חזית החנות"],
        "about": "היא מכולת שכונתית שבה עדיין מכירים אותך בשם. אנחנו דואגים שהמוצרים יהיו טריים, המחירים הוגנים, והשירות אישי — בדיוק כמו שאוהבים.",
    },
    "falafel": {
        "hero_emoji": "🧆",
        "nav": {"services": "התפריט", "gallery": "מהמטבח", "hours": "שעות", "contact": "הזמנות"},
        "services_title": "התפריט שלנו",
        "services_note": "התפריט להמחשה — הכול מוכן טרי במקום.",
        "services": [
            ("🧆", "מנת פלאפל", "פלאפל טרי וקריספי בפיתה או לאפה, עם סלטים."),
            ("🥙", "חומוס ביתי", "חומוס חלק עם גרגירים, טחינה ושמן זית."),
            ("🍽️", "מסבחה ופול", "מסבחה חמה ופול ביתי — כמו שצריך."),
            ("🥗", "מגש סלטים", "מבחר סלטים טריים שמתחלפים מדי יום."),
            ("🍟", "תוספות", "צ'יפס טרי, חצילים ותוספות לבחירה."),
            ("🥤", "משקאות", "שתייה קרה, לימונדה ביתית ושתייה חמה."),
        ],
        "features": [
            ("🔥", "טרי במקום", "מטגנים ומכינים הכול לפי הזמנה."),
            ("⚡", "מהיר וטעים", "מנה משביעה תוך דקות, גם בהפסקת צהריים."),
            ("🌿", "חומרי גלם איכותיים", "מתכונים ביתיים וחומרי גלם נבחרים."),
        ],
        "gallery": ["פלאפל טרי", "צלחת חומוס", "מסבחה חמה", "מגש הסלטים", "פיתה במילוי", "הדלפק שלנו"],
        "about": "הוא מקום קטן עם טעם גדול. אנחנו מכינים חומוס, פלאפל ואוכל ביתי טרי כל יום, עם חומרי גלם איכותיים ובאהבה — מנה שמרגישה כמו בית.",
    },
    "florist": {
        "hero_emoji": "💐",
        "nav": {"services": "הזרים שלנו", "gallery": "גלריה", "hours": "שעות", "contact": "הזמנות"},
        "services_title": "מה אנחנו מכינים",
        "services_note": "כל זר נעשה בעבודת יד — אפשר להתאים אישית.",
        "services": [
            ("💐", "זרים לכל אירוע", "זרים מעוצבים בעבודת יד לפי סגנון וצבע."),
            ("🌹", "זרי אהבה", "זרי ורדים ומעמדים לרגעים המיוחדים."),
            ("🪴", "צמחים ועציצים", "צמחי בית ומתנות ירוקות שנשארות לאורך זמן."),
            ("🎁", "מארזי מתנה", "שילובים של פרחים, שוקולד ומתוקים."),
            ("💒", "אירועים וחתונות", "עיצוב פרחים לאירועים קטנים וגדולים."),
            ("🚚", "משלוחים לגוש דן", "משלוח פרחים טריים עד הבית או המשרד."),
        ],
        "features": [
            ("🌸", "טריות מובטחת", "פרחים טריים שנבחרים בקפידה כל בוקר."),
            ("🎨", "עיצוב אישי", "מתאימים את הזר בדיוק לאירוע ולתקציב."),
            ("🚚", "משלוח מהיר", "משלוחים לגבעתיים, תל אביב ורמת גן."),
        ],
        "gallery": ["זר צבעוני", "זר ורדים", "עציץ מעוצב", "מארז מתנה", "סידור לאירוע", "בחנות"],
        "about": "היא חנות פרחים שמאמינה שכל זר מספר סיפור. אנחנו מעצבים בעבודת יד, עם פרחים טריים ותשומת לב לפרטים, ומשלימים כל רגע מיוחד עם משלוח לכל גוש דן.",
    },
    "cobbler": {
        "hero_emoji": "👞",
        "nav": {"services": "השירותים", "gallery": "העבודות", "hours": "שעות", "contact": "צור קשר"},
        "services_title": "מה אנחנו מתקנים",
        "services_note": "מחזירים לחיים כל פריט — מהיום להיום.",
        "services": [
            ("👞", "תיקון נעליים", "החלפת סוליות, עקבים ותפירה של כל סוגי הנעליים."),
            ("👜", "תיקון תיקים וחגורות", "תיקוני עור, רוכסנים וקצוות לתיקים וחגורות."),
            ("🧵", "תיקוני בגדים", "הצרות, הרחבות ומכפלות בהתאמה מדויקת."),
            ("🔑", "שכפול מפתחות", "העתקת מפתחות לבית ולעסק במקום."),
            ("🧥", "מעילים ומעילי עור", "טיפול, תיקון ותפירה של מעילים ובגדי עור."),
            ("✨", "צחצוח וטיפוח עור", "ניקוי, הזנה וצביעה שמחזירים מראה חדש."),
        ],
        "features": [
            ("⏱️", "שירות מהיר", "רוב התיקונים מוכנים כבר באותו היום."),
            ("🛠️", "מקצועיות ותיקה", "שנים של ניסיון בעבודת יד איכותית."),
            ("💬", "ייעוץ כן", "אומרים לך בכנות אם משתלם לתקן."),
        ],
        "gallery": ["החלפת סוליה", "תיקון עקב", "תפירת עור", "תיקון תיק", "מכפלת מכנס", "השולחן שלנו"],
        "about": "הוא מקום שבו יודעים לתקן במקום לזרוק. בעבודת יד, בסבלנות ובמקצועיות ותיקה, אנחנו מחזירים לנעליים, לתיקים ולבגדים חיים חדשים — מהר ובאיכות.",
    },
    "cosmetics": {
        "hero_emoji": "💅",
        "nav": {"services": "הטיפולים", "gallery": "גלריה", "hours": "שעות", "contact": "קביעת תור"},
        "services_title": "הטיפולים שלנו",
        "services_note": "כל טיפול מותאם אישית — נשמח לייעץ.",
        "services": [
            ("💅", "מניקור וג'ל", "מניקור, לק ג'ל ובנייה בגימור מושלם."),
            ("🦶", "פדיקור", "פדיקור קוסמטי וטיפוח כפות הרגליים."),
            ("✨", "טיפולי פנים", "ניקוי עור, הזנה וטיפולים מותאמים לעור."),
            ("🌿", "הסרת שיער בשעווה", "הסרת שיער עדינה ויסודית בכל אזורי הגוף."),
            ("👁️", "עיצוב גבות וריסים", "עיצוב, שיפוד וצביעה של גבות וריסים."),
            ("💆", "טיפולי טיפוח", "פינוק, מסכות וטיפולים לעור זוהר."),
        ],
        "features": [
            ("💗", "יחס אישי וחם", "מקשיבים לך ומתאימים את הטיפול בדיוק לך."),
            ("🧴", "חומרים איכותיים", "מוצרים מקצועיים ואיכותיים בלבד."),
            ("🕯️", "אווירה נעימה", "חוויית פינוק רגועה מהרגע שנכנסים."),
        ],
        "gallery": ["מניקור ג'ל", "עיצוב ציפורניים", "טיפול פנים", "עיצוב גבות", "פינת הטיפולים", "התוצאה"],
        "about": "הוא בוטיק טיפוח אישי שבו כל לקוחה מרגישה מיוחדת. עם חומרים איכותיים, יד מקצועית ואווירה נעימה, אנחנו דואגים שתצאו מכאן זוהרות ומחויכות.",
    },
    "bakery": {
        "hero_emoji": "🥐",
        "nav": {"services": "המאפים", "gallery": "מהתנור", "hours": "שעות", "contact": "הזמנות"},
        "services_title": "מה יוצא מהתנור",
        "services_note": "נאפה טרי כל היום — התפריט להמחשה.",
        "services": [
            ("🥟", "בורקס במגוון מילויים", "בורקס גבינה, תפוחי אדמה, פיצה ועוד — חם מהתנור."),
            ("🥐", "מאפים מלוחים", "ג'חנון, סיגרים, מלאווח ומאפים חמים."),
            ("🍰", "מאפים מתוקים", "עוגיות, שטרודל, בורקס מתוק ומאפי שוקולד."),
            ("🍞", "לחם ולחמניות", "לחם טרי ולחמניות שנאפים כל בוקר."),
            ("🥧", "מגשי אירוח", "מגשים למשרד, אירוע או שבת — לפי הזמנה."),
            ("☕", "פינת קפה", "קפה טרי לצד המאפה, לשבת או לדרך."),
        ],
        "features": [
            ("🔥", "טרי מהתנור", "אופים לאורך כל היום — תמיד חם."),
            ("👐", "עבודת יד", "מתכונים מסורתיים בעבודת יד אמיתית."),
            ("📦", "מגשים בהזמנה", "מזמינים מראש מגשי אירוח לכל אירוע."),
        ],
        "gallery": ["בורקס חם", "מגש מאפים", "ג'חנון", "מאפה שוקולד", "לחם טרי", "התנור שלנו"],
        "about": "היא מאפייה שכונתית שבה אופים טרי לאורך כל היום. עם מתכונים מסורתיים ועבודת יד, אנחנו מוציאים מהתנור בורקס ומאפים חמים שמריחים אותם כבר מהרחוב.",
    },
    "tailor": {
        "hero_emoji": "🧵",
        "nav": {"services": "השירותים", "gallery": "העבודות", "hours": "שעות", "contact": "צור קשר"},
        "services_title": "השירותים שלנו",
        "services_note": "כל תיקון בהתאמה מדויקת — עבודת יד.",
        "services": [
            ("📏", "הצרות והרחבות", "התאמת מידה מדויקת לכל פריט לבוש."),
            ("✂️", "מכפלות וקיצור", "קיצור מכנסיים, חצאיות ושמלות בגימור נקי."),
            ("🤵", "תיקוני חליפות", "התאמה ותיקון של חליפות וז'קטים."),
            ("👰", "שמלות ערב וכלה", "התאמות עדינות לשמלות אירוע וכלה."),
            ("🔻", "החלפת רוכסנים", "החלפה ותיקון של רוכסנים וסגירות."),
            ("🧶", "תיקונים ושחזור", "תיקון קרעים, כפתורים ושחזור בגדים אהובים."),
        ],
        "features": [
            ("🎯", "התאמה מושלמת", "עין מקצועית לכל תפר ומידה."),
            ("⏱️", "בזמן שהובטח", "עומדים בלוחות זמנים, גם לאירועים."),
            ("💎", "גימור איכותי", "עבודת יד נקייה שמחזיקה לאורך זמן."),
        ],
        "gallery": ["התאמת חליפה", "מכפלת שמלה", "החלפת רוכסן", "תיקון עדין", "פינת התפירה", "פרטים"],
        "about": "היא מתפרה שבה כל בגד מקבל יחס אישי. עם עין מקצועית ועבודת יד מדויקת, אנחנו מתאימים, מתקנים ומשחזרים — כדי שכל פריט ישב עליך בדיוק כמו שצריך.",
    },
}

# פרסונות גנריות לביקורות (ללא שמות אמיתיים — טקסט לדוגמה שהעסק יחליף)
REVIEWS = [
    ("לקוח/ה קבוע/ה", "שירות אדיב ומקצועי, תמיד יוצאים מרוצים. ממליץ בחום!"),
    ("תושב/ת השכונה", "יחס אישי וחם, מרגישים כמו במשפחה. עסק שנעים לתמוך בו."),
    ("לקוח/ה חדש/ה", "הגעתי בהמלצה ונשארתי. איכות ומחיר הוגן, בדיוק מה שחיפשתי."),
]

# ----------------------------------------------------------------------------
# עוזרים
# ----------------------------------------------------------------------------
def esc(s):
    return html.escape(s or "", quote=True)

def digits(phone):
    return re.sub(r"\D", "", phone or "")

def whatsapp_link(phone):
    d = digits(phone)
    if not d:
        return None
    if d.startswith("0"):
        d = "972" + d[1:]
    return "https://wa.me/" + d

def maps_link(address):
    return "https://www.google.com/maps/search/?api=1&query=" + \
        re.sub(r"\s+", "+", address)

def waze_link(address):
    return "https://waze.com/ul?q=" + re.sub(r"\s+", "%20", address)

def tel_link(phone):
    d = digits(phone)
    return "tel:" + d if d else None

YEAR = datetime.date.today().year

# ----------------------------------------------------------------------------
# תבנית אתר בודד
# ----------------------------------------------------------------------------
def render_site(b):
    cat = CATEGORIES[b["category"]]
    pal = PALETTES.get(b.get("theme", "slate"), PALETTES["slate"])
    name = b["name"]
    short = b.get("short_name", name)
    nav = cat["nav"]
    phone = b.get("phone")
    tel = tel_link(phone)
    wa = whatsapp_link(phone)
    maps = maps_link(b["address"])
    waze = waze_link(b["address"])

    # כפתורי CTA בהתאם למידע הזמין
    hero_ctas = []
    if tel:
        hero_ctas.append(f'<a class="btn btn-primary" href="{tel}">📞 התקשרו: {esc(phone)}</a>')
    if wa:
        hero_ctas.append(f'<a class="btn btn-ghost" href="{wa}" target="_blank" rel="noopener">💬 וואטסאפ</a>')
    hero_ctas.append(f'<a class="btn btn-ghost" href="{maps}" target="_blank" rel="noopener">📍 ניווט</a>')
    hero_ctas_html = "\n            ".join(hero_ctas)

    # שירותים
    services_html = "\n".join(
        f'''          <li class="card">
            <span class="card-emoji" aria-hidden="true">{s[0]}</span>
            <h3>{esc(s[1])}</h3>
            <p>{esc(s[2])}</p>
          </li>''' for s in cat["services"])

    # יתרונות
    features_html = "\n".join(
        f'''          <li class="feature">
            <span class="feature-emoji" aria-hidden="true">{f[0]}</span>
            <div><h3>{esc(f[1])}</h3><p>{esc(f[2])}</p></div>
          </li>''' for f in cat["features"])

    # גלריה (כרטיסי תמונה מבוססי גרדיאנט + אמוג'י, ללא תלות בקבצים חיצוניים)
    gal_emoji = cat["hero_emoji"]
    gallery_html = "\n".join(
        f'''          <figure class="shot shot-{i%3}">
            <span class="shot-emoji" aria-hidden="true">{gal_emoji}</span>
            <figcaption>{esc(label)}</figcaption>
          </figure>''' for i, label in enumerate(cat["gallery"]))

    # ביקורות
    reviews_html = "\n".join(
        f'''          <blockquote class="review">
            <p>“{esc(txt)}”</p>
            <cite>— {esc(who)}</cite>
          </blockquote>''' for who, txt in REVIEWS)

    # פרטי קשר
    contact_rows = [
        f'<div class="ci"><span aria-hidden="true">📍</span><div><dt>כתובת</dt><dd><a href="{maps}" target="_blank" rel="noopener">{esc(b["address"])}</a></dd></div></div>',
    ]
    if phone:
        contact_rows.append(f'<div class="ci"><span aria-hidden="true">📞</span><div><dt>טלפון</dt><dd><a href="{tel}">{esc(phone)}</a></dd></div></div>')
    if wa:
        contact_rows.append(f'<div class="ci"><span aria-hidden="true">💬</span><div><dt>וואטסאפ</dt><dd><a href="{wa}" target="_blank" rel="noopener">שליחת הודעה</a></dd></div></div>')
    contact_rows.append(f'<div class="ci"><span aria-hidden="true">🧭</span><div><dt>ניווט</dt><dd><a href="{waze}" target="_blank" rel="noopener">Waze</a> · <a href="{maps}" target="_blank" rel="noopener">Google Maps</a></dd></div></div>')
    contact_html = "\n            ".join(contact_rows)

    about_full = f"{esc(short)} {esc(cat['about'])}"

    return f'''<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(name)} | {esc(b["category_he"])} ב{esc(b["city"])}</title>
<meta name="description" content="{esc(name)} — {esc(b['tagline'])}. {esc(b['category_he'])} ב{esc(b['city'])}, {esc(b['address'])}.">
<meta name="theme-color" content="{pal['primary']}">
<meta property="og:title" content="{esc(name)}">
<meta property="og:description" content="{esc(b['tagline'])}">
<meta property="og:type" content="website">
<script type="application/ld+json">
{json.dumps({
  "@context": "https://schema.org",
  "@type": "LocalBusiness",
  "name": name,
  "description": b['tagline'],
  "address": {"@type": "PostalAddress", "streetAddress": b['address'], "addressLocality": b['city'], "addressCountry": "IL"},
  **({"telephone": phone} if phone else {}),
  "areaServed": b['city'],
}, ensure_ascii=False)}
</script>
<style>
:root{{
  --primary:{pal['primary']}; --dark:{pal['dark']}; --accent:{pal['accent']}; --ink:{pal['ink']};
  --bg:#ffffff; --surface:#f7f7f9; --text:#1a1a1e; --muted:#5b5b66; --line:#e6e6ec;
  --radius:16px; --maxw:1080px; --shadow:0 10px 30px rgba(0,0,0,.08);
  --font:'Segoe UI',system-ui,'Assistant','Rubik','Heebo',Arial,sans-serif;
}}
*{{box-sizing:border-box}}
html{{scroll-behavior:smooth}}
body{{margin:0;font-family:var(--font);color:var(--text);background:var(--bg);line-height:1.7;font-size:17px}}
h1,h2,h3{{line-height:1.25;margin:0 0 .4em}}
h2{{font-size:clamp(1.5rem,3.5vw,2.1rem)}}
p{{margin:0 0 1em}}
a{{color:var(--primary)}}
img{{max-width:100%}}
.container{{max-width:var(--maxw);margin-inline:auto;padding-inline:20px}}
.skip{{position:absolute;right:-999px;top:0;background:var(--primary);color:#fff;padding:10px 16px;z-index:100;border-radius:0 0 10px 10px}}
.skip:focus{{right:12px}}
:focus-visible{{outline:3px solid var(--accent);outline-offset:2px;border-radius:6px}}

/* Header */
.site-header{{position:sticky;top:0;z-index:50;background:rgba(255,255,255,.92);backdrop-filter:blur(8px);border-bottom:1px solid var(--line)}}
.nav{{display:flex;align-items:center;justify-content:space-between;gap:16px;padding-block:12px}}
.brand{{display:flex;align-items:center;gap:10px;font-weight:800;font-size:1.15rem;color:var(--dark);text-decoration:none}}
.brand .logo{{width:40px;height:40px;border-radius:12px;display:grid;place-items:center;background:linear-gradient(135deg,var(--primary),var(--dark));color:#fff;font-size:1.2rem}}
.nav-links{{display:flex;gap:6px;align-items:center;list-style:none;margin:0;padding:0}}
.nav-links a{{text-decoration:none;color:var(--text);padding:8px 12px;border-radius:10px;font-weight:600}}
.nav-links a:hover{{background:var(--surface)}}
.nav-cta{{background:var(--primary);color:#fff!important}}
.nav-toggle{{display:none;background:none;border:1px solid var(--line);border-radius:10px;padding:8px 12px;font-size:1.2rem;cursor:pointer}}

/* Hero */
.hero{{position:relative;color:#fff;background:linear-gradient(135deg,var(--primary),var(--dark));overflow:hidden}}
.hero::before{{content:"";position:absolute;inset:0;background-image:radial-gradient(circle at 20% 20%,rgba(255,255,255,.14),transparent 40%),radial-gradient(circle at 85% 80%,rgba(255,255,255,.1),transparent 40%);pointer-events:none}}
.hero-inner{{position:relative;display:grid;grid-template-columns:1.3fr .7fr;gap:30px;align-items:center;padding-block:64px}}
.hero .eyebrow{{display:inline-block;background:rgba(255,255,255,.16);padding:6px 14px;border-radius:999px;font-weight:700;font-size:.9rem;margin-bottom:14px}}
.hero h1{{font-size:clamp(2rem,5vw,3.2rem);color:#fff}}
.hero .lead{{font-size:1.2rem;color:var(--ink);max-width:44ch}}
.hero-badge{{aspect-ratio:1;display:grid;place-items:center;font-size:clamp(4rem,14vw,8rem);background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.25);border-radius:28px}}
.cta-row{{display:flex;flex-wrap:wrap;gap:12px;margin-top:22px}}
.btn{{display:inline-flex;align-items:center;gap:8px;padding:13px 22px;border-radius:12px;font-weight:800;text-decoration:none;border:2px solid transparent;cursor:pointer;font-size:1rem}}
.btn-primary{{background:#fff;color:var(--dark)}}
.btn-primary:hover{{transform:translateY(-1px)}}
.btn-ghost{{background:rgba(255,255,255,.1);color:#fff;border-color:rgba(255,255,255,.55)}}
.btn-ghost:hover{{background:rgba(255,255,255,.2)}}

/* Sections */
section{{padding-block:60px}}
.section-head{{text-align:center;max-width:60ch;margin-inline:auto;margin-bottom:36px}}
.section-head .kicker{{color:var(--primary);font-weight:800;letter-spacing:.04em;text-transform:uppercase;font-size:.85rem}}
.note{{color:var(--muted);font-size:.95rem}}

.grid{{display:grid;gap:18px;list-style:none;margin:0;padding:0}}
.services .grid{{grid-template-columns:repeat(3,1fr)}}
.card{{background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:24px;transition:transform .15s,box-shadow .15s}}
.card:hover{{transform:translateY(-3px);box-shadow:var(--shadow)}}
.card-emoji{{font-size:2rem;display:inline-block;margin-bottom:10px}}
.card h3{{font-size:1.15rem}}
.card p{{color:var(--muted);margin:0}}

.features{{background:var(--surface)}}
.features .grid{{grid-template-columns:repeat(3,1fr)}}
.feature{{display:flex;gap:14px;align-items:flex-start;background:#fff;border:1px solid var(--line);border-radius:var(--radius);padding:22px}}
.feature-emoji{{font-size:1.8rem}}
.feature h3{{font-size:1.1rem;margin:0 0 .2em}}
.feature p{{color:var(--muted);margin:0}}

.gallery .grid{{grid-template-columns:repeat(3,1fr)}}
.shot{{margin:0;border-radius:var(--radius);overflow:hidden;position:relative;aspect-ratio:4/3;display:grid;place-items:center;color:#fff}}
.shot-0{{background:linear-gradient(135deg,var(--primary),var(--dark))}}
.shot-1{{background:linear-gradient(135deg,var(--dark),var(--primary))}}
.shot-2{{background:linear-gradient(135deg,var(--accent),var(--primary))}}
.shot-emoji{{font-size:3.4rem;filter:drop-shadow(0 4px 8px rgba(0,0,0,.25))}}
.shot figcaption{{position:absolute;inset-inline:0;bottom:0;background:linear-gradient(transparent,rgba(0,0,0,.55));padding:14px;font-weight:700}}

.about-wrap{{display:grid;grid-template-columns:1fr 1fr;gap:36px;align-items:center}}
.about-card{{background:linear-gradient(135deg,var(--primary),var(--dark));color:#fff;border-radius:24px;padding:36px;min-height:220px;display:grid;place-content:center;text-align:center}}
.about-card .big{{font-size:5rem}}

.reviews .grid{{grid-template-columns:repeat(3,1fr)}}
.review{{background:var(--surface);border:1px solid var(--line);border-radius:var(--radius);padding:24px;margin:0}}
.review p{{font-size:1.05rem}}
.review cite{{color:var(--primary);font-weight:700;font-style:normal}}

.hours-contact{{background:var(--surface)}}
.hc-grid{{display:grid;grid-template-columns:1fr 1fr;gap:30px}}
.panel{{background:#fff;border:1px solid var(--line);border-radius:var(--radius);padding:28px}}
.hours-list{{list-style:none;margin:0;padding:0}}
.hours-list li{{display:flex;justify-content:space-between;padding:10px 0;border-bottom:1px dashed var(--line)}}
.hours-list li:last-child{{border-bottom:0}}
.hours-list .day{{font-weight:700}}
.contact-info{{margin:0}}
.ci{{display:flex;gap:12px;align-items:flex-start;padding:12px 0;border-bottom:1px dashed var(--line)}}
.ci:last-child{{border-bottom:0}}
.ci span{{font-size:1.3rem}}
.ci dt{{font-weight:800;font-size:.85rem;color:var(--muted)}}
.ci dd{{margin:0;font-size:1.05rem}}

form{{display:grid;gap:14px;margin-top:18px}}
.field{{display:grid;gap:6px}}
label{{font-weight:700;font-size:.95rem}}
input,textarea{{font:inherit;padding:12px 14px;border:1.5px solid var(--line);border-radius:10px;background:#fff;width:100%}}
input:focus,textarea:focus{{border-color:var(--primary);outline:none}}
.form-msg{{padding:12px 14px;border-radius:10px;font-weight:700;display:none}}
.form-msg.ok{{display:block;background:#ecfdf5;color:#065f46;border:1px solid #6ee7b7}}
.form-msg.err{{display:block;background:#fef2f2;color:#991b1b;border:1px solid #fca5a5}}

.site-footer{{background:var(--dark);color:#fff;padding-block:40px}}
.site-footer a{{color:#fff}}
.foot-grid{{display:grid;grid-template-columns:1.4fr 1fr 1fr;gap:26px}}
.foot-grid h4{{margin:0 0 12px;font-size:1rem}}
.foot-grid ul{{list-style:none;margin:0;padding:0;display:grid;gap:8px}}
.foot-note{{border-top:1px solid rgba(255,255,255,.15);margin-top:28px;padding-top:18px;font-size:.85rem;color:#cbd5e1;display:flex;justify-content:space-between;flex-wrap:wrap;gap:10px}}

@media (max-width:860px){{
  .hero-inner,.about-wrap,.hc-grid{{grid-template-columns:1fr}}
  .hero-badge{{display:none}}
  .services .grid,.features .grid,.gallery .grid,.reviews .grid,.foot-grid{{grid-template-columns:1fr 1fr}}
  .nav-links{{position:fixed;inset:64px 0 auto 0;flex-direction:column;background:#fff;border-bottom:1px solid var(--line);padding:14px 20px;display:none}}
  .nav-links.open{{display:flex}}
  .nav-links a{{width:100%}}
  .nav-toggle{{display:inline-block}}
}}
@media (max-width:520px){{
  .services .grid,.features .grid,.gallery .grid,.reviews .grid,.foot-grid{{grid-template-columns:1fr}}
}}
@media (prefers-reduced-motion:reduce){{
  *{{scroll-behavior:auto!important;transition:none!important}}
}}
</style>
</head>
<body>
<a class="skip" href="#main">דלג לתוכן</a>

<header class="site-header">
  <div class="container nav">
    <a class="brand" href="#top"><span class="logo" aria-hidden="true">{cat['hero_emoji']}</span><span>{esc(short)}</span></a>
    <button class="nav-toggle" aria-expanded="false" aria-controls="menu" aria-label="פתיחת תפריט">☰</button>
    <ul class="nav-links" id="menu">
      <li><a href="#services">{esc(nav['services'])}</a></li>
      <li><a href="#gallery">{esc(nav['gallery'])}</a></li>
      <li><a href="#hours">{esc(nav['hours'])}</a></li>
      <li><a class="nav-cta" href="#contact">{esc(nav['contact'])}</a></li>
    </ul>
  </div>
</header>

<main id="main">
<section class="hero" id="top" aria-labelledby="hero-title">
  <div class="container hero-inner">
    <div>
      <span class="eyebrow">{esc(b['category_he'])} · {esc(b['city'])}</span>
      <h1 id="hero-title">{esc(name)}</h1>
      <p class="lead">{esc(b['tagline'])}</p>
      <div class="cta-row">
            {hero_ctas_html}
      </div>
    </div>
    <div class="hero-badge" aria-hidden="true">{cat['hero_emoji']}</div>
  </div>
</section>

<section class="services" id="services" aria-labelledby="services-title">
  <div class="container">
    <div class="section-head">
      <span class="kicker">{esc(b['category_he'])}</span>
      <h2 id="services-title">{esc(cat['services_title'])}</h2>
      <p class="note">{esc(cat['services_note'])}</p>
    </div>
    <ul class="grid">
{services_html}
    </ul>
  </div>
</section>

<section class="features" aria-labelledby="features-title">
  <div class="container">
    <div class="section-head">
      <span class="kicker">למה אנחנו</span>
      <h2 id="features-title">מה שאתם מקבלים אצלנו</h2>
    </div>
    <ul class="grid">
{features_html}
    </ul>
  </div>
</section>

<section class="about" aria-labelledby="about-title">
  <div class="container about-wrap">
    <div>
      <span class="kicker" style="color:var(--primary);font-weight:800">קצת עלינו</span>
      <h2 id="about-title">הסיפור של {esc(short)}</h2>
      <p>{about_full}</p>
      <p>אנחנו נמצאים ב{esc(b['address'])} ומשרתים את {esc(b['city'])} והסביבה. מוזמנים לקפוץ, להתקשר או לשלוח הודעה — תמיד נשמח לעזור.</p>
    </div>
    <div class="about-card">
      <div class="big" aria-hidden="true">{cat['hero_emoji']}</div>
      <p style="margin:0;font-weight:700;font-size:1.2rem">{esc(b['city'])}</p>
      <p style="margin:0;color:var(--ink)">{esc(b['address'])}</p>
    </div>
  </div>
</section>

<section class="gallery" id="gallery" aria-labelledby="gallery-title">
  <div class="container">
    <div class="section-head">
      <span class="kicker">{esc(nav['gallery'])}</span>
      <h2 id="gallery-title">רגעים מאצלנו</h2>
      <p class="note">התמונות להמחשה — אפשר להחליף בתמונות אמיתיות של העסק.</p>
    </div>
    <div class="grid">
{gallery_html}
    </div>
  </div>
</section>

<section class="reviews" aria-labelledby="reviews-title">
  <div class="container">
    <div class="section-head">
      <span class="kicker">המלצות</span>
      <h2 id="reviews-title">מה אומרים עלינו</h2>
      <p class="note">ביקורות לדוגמה — מומלץ להחליף בהמלצות אמיתיות של לקוחות.</p>
    </div>
    <div class="grid">
{reviews_html}
    </div>
  </div>
</section>

<section class="hours-contact" id="contact" aria-labelledby="contact-title">
  <div class="container">
    <div class="section-head">
      <span class="kicker">{esc(nav['contact'])}</span>
      <h2 id="contact-title">בואו נדבר</h2>
    </div>
    <div class="hc-grid">
      <div class="panel" id="hours">
        <h3>שעות פעילות</h3>
        <ul class="hours-list">
          <li><span class="day">ראשון–חמישי</span><span>09:00–19:00</span></li>
          <li><span class="day">שישי / ערבי חג</span><span>09:00–14:00</span></li>
          <li><span class="day">שבת</span><span>סגור</span></li>
        </ul>
        <p class="note" style="margin-top:14px">שעות הפעילות להמחשה — ניתן לעדכן לשעות המדויקות של העסק.</p>
        <h3 style="margin-top:24px">פרטים ליצירת קשר</h3>
        <dl class="contact-info">
            {contact_html}
        </dl>
      </div>
      <div class="panel">
        <h3>השאירו הודעה</h3>
        <p class="note">מלאו את הפרטים ונחזור אליכם בהקדם.</p>
        <form id="contact-form" novalidate>
          <div class="field">
            <label for="cf-name">שם מלא</label>
            <input id="cf-name" name="name" type="text" autocomplete="name" required aria-required="true">
          </div>
          <div class="field">
            <label for="cf-phone">טלפון</label>
            <input id="cf-phone" name="phone" type="tel" inputmode="tel" autocomplete="tel" required aria-required="true">
          </div>
          <div class="field">
            <label for="cf-msg">הודעה</label>
            <textarea id="cf-msg" name="message" rows="4"></textarea>
          </div>
          <p class="form-msg" id="cf-status" role="status" aria-live="polite"></p>
          <button class="btn btn-primary" type="submit" style="background:var(--primary);color:#fff">שליחה</button>
        </form>
      </div>
    </div>
  </div>
</section>
</main>

<footer class="site-footer">
  <div class="container">
    <div class="foot-grid">
      <div>
        <h4>{esc(name)}</h4>
        <p style="color:#cbd5e1">{esc(b['tagline'])}</p>
        <p style="color:#cbd5e1">{esc(b['category_he'])} · {esc(b['city'])}</p>
      </div>
      <div>
        <h4>ניווט</h4>
        <ul>
          <li><a href="#services">{esc(nav['services'])}</a></li>
          <li><a href="#gallery">{esc(nav['gallery'])}</a></li>
          <li><a href="#hours">{esc(nav['hours'])}</a></li>
          <li><a href="#contact">{esc(nav['contact'])}</a></li>
        </ul>
      </div>
      <div>
        <h4>יצירת קשר</h4>
        <ul>
          <li><a href="{maps}" target="_blank" rel="noopener">{esc(b['address'])}</a></li>
          {f'<li><a href="{tel}">{esc(phone)}</a></li>' if phone else ''}
          {f'<li><a href="{wa}" target="_blank" rel="noopener">וואטסאפ</a></li>' if wa else ''}
        </ul>
      </div>
    </div>
    <div class="foot-note">
      <span>© {YEAR} {esc(name)}. כל הזכויות שמורות.</span>
      <span>אתר לדוגמה שנבנה עבור העסק · התוכן ניתן לעדכון</span>
    </div>
  </div>
</footer>

<script>
(function(){{
  var toggle=document.querySelector('.nav-toggle'),menu=document.getElementById('menu');
  if(toggle){{toggle.addEventListener('click',function(){{
    var open=menu.classList.toggle('open');
    toggle.setAttribute('aria-expanded',open?'true':'false');
  }});
  menu.addEventListener('click',function(e){{if(e.target.tagName==='A'){{menu.classList.remove('open');toggle.setAttribute('aria-expanded','false');}}}});}}

  var form=document.getElementById('contact-form'),status=document.getElementById('cf-status');
  if(form){{form.addEventListener('submit',function(e){{
    e.preventDefault();
    var name=form.name.value.trim(),phone=form.phone.value.trim();
    status.className='form-msg';
    if(!name||!phone){{status.classList.add('err');status.textContent='נא למלא שם וטלפון.';return;}}
    if(!/^[0-9+\\-\\s()]{{7,}}$/.test(phone)){{status.classList.add('err');status.textContent='מספר הטלפון אינו תקין.';return;}}
    status.classList.add('ok');
    status.textContent='תודה '+name+'! ההודעה התקבלה ונחזור אליך בהקדם.';
    form.reset();
  }});}}
}})();
</script>
</body>
</html>
'''


# ----------------------------------------------------------------------------
# עמוד אינדקס
# ----------------------------------------------------------------------------
def render_index(businesses):
    cards = []
    for b in businesses:
        cat = CATEGORIES[b["category"]]
        pal = PALETTES.get(b.get("theme", "slate"), PALETTES["slate"])
        phone = b.get("phone") or "לאימות"
        cards.append(f'''      <a class="tile" href="./{b['slug']}/index.html" style="--c:{pal['primary']};--d:{pal['dark']}">
        <span class="tile-emoji" aria-hidden="true">{cat['hero_emoji']}</span>
        <span class="tile-cat">{esc(b['category_he'])}</span>
        <h2>{esc(b['name'])}</h2>
        <p class="tile-loc">📍 {esc(b['address'])}</p>
        <p class="tile-phone">📞 {esc(phone)}</p>
        <span class="tile-go">לאתר ←</span>
      </a>''')
    cards_html = "\n".join(cards)
    n = len(businesses)
    cities = sorted(set(b["city"] for b in businesses))
    return f'''<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>אתרים לעסקים מקומיים · גבעתיים · תל אביב · רמת גן</title>
<meta name="description" content="{n} אתרים מותאמים אישית לעסקים מקומיים ללא אתר, באזור גוש דן.">
<style>
*{{box-sizing:border-box}}
body{{margin:0;font-family:'Segoe UI',system-ui,'Assistant','Rubik',Arial,sans-serif;background:#0f172a;color:#e2e8f0;line-height:1.7}}
a{{color:inherit;text-decoration:none}}
.wrap{{max-width:1120px;margin-inline:auto;padding:0 20px}}
header{{padding:64px 0 40px;text-align:center;background:radial-gradient(circle at 50% 0%,#1e293b,#0f172a)}}
header .eyebrow{{display:inline-block;background:#1e293b;border:1px solid #334155;padding:6px 16px;border-radius:999px;font-weight:700;color:#93c5fd;margin-bottom:16px}}
header h1{{font-size:clamp(1.9rem,5vw,3rem);margin:0 0 12px}}
header p{{color:#94a3b8;max-width:60ch;margin:0 auto;font-size:1.1rem}}
.stats{{display:flex;gap:28px;justify-content:center;flex-wrap:wrap;margin-top:26px}}
.stat b{{display:block;font-size:2rem;color:#fff}}
.stat span{{color:#94a3b8;font-size:.9rem}}
main{{padding:40px 0 70px}}
.grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:20px}}
.tile{{background:linear-gradient(160deg,#1e293b,#0f172a);border:1px solid #334155;border-radius:18px;padding:26px;display:flex;flex-direction:column;gap:6px;position:relative;overflow:hidden;transition:transform .15s,border-color .15s}}
.tile::before{{content:"";position:absolute;inset:0 0 auto 0;height:5px;background:linear-gradient(90deg,var(--c),var(--d))}}
.tile:hover{{transform:translateY(-4px);border-color:var(--c)}}
.tile:focus-visible{{outline:3px solid var(--c);outline-offset:3px}}
.tile-emoji{{font-size:2.4rem}}
.tile-cat{{color:#93c5fd;font-weight:700;font-size:.85rem}}
.tile h2{{font-size:1.2rem;margin:.1em 0}}
.tile-loc,.tile-phone{{color:#94a3b8;margin:0;font-size:.92rem}}
.tile-go{{margin-top:10px;font-weight:800;color:#fff}}
footer{{border-top:1px solid #1e293b;padding:30px 0;text-align:center;color:#64748b;font-size:.9rem}}
@media (max-width:860px){{.grid{{grid-template-columns:1fr 1fr}}}}
@media (max-width:560px){{.grid{{grid-template-columns:1fr}}}}
@media (prefers-reduced-motion:reduce){{*{{transition:none!important}}}}
</style>
</head>
<body>
<header>
  <div class="wrap">
    <span class="eyebrow">גוש דן · {esc(' · '.join(cities))}</span>
    <h1>אתרים לעסקים מקומיים</h1>
    <p>{n} אתרים מלאים, נגישים ומותאמים אישית — שנבנו לעסקים אמיתיים באזור שאין להם עדיין אתר אינטרנט.</p>
    <div class="stats">
      <div class="stat"><b>{n}</b><span>אתרים</span></div>
      <div class="stat"><b>{len(cities)}</b><span>ערים</span></div>
      <div class="stat"><b>{len(set(b['category'] for b in businesses))}</b><span>תחומים</span></div>
    </div>
  </div>
</header>
<main>
  <div class="wrap">
    <div class="grid">
{cards_html}
    </div>
  </div>
</main>
<footer>
  <div class="wrap">נבנה כפרויקט הדגמה · כל הפרטים נאספו ממקורות פומביים ומיועדים לאימות לפני פנייה לעסק.</div>
</footer>
</body>
</html>
'''


# ----------------------------------------------------------------------------
# קובץ אקסל
# ----------------------------------------------------------------------------
def build_excel(businesses, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "לידים"
    ws.sheet_view.rightToLeft = True

    headers = ["#", "שם העסק", "תחום", "עיר", "כתובת", "טלפון",
               "וואטסאפ", "סטטוס אתר", "מקור", "קישור למקור", "אתר שנבנה"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for i, b in enumerate(businesses, 1):
        wa = whatsapp_link(b.get("phone")) or ""
        row = [
            i, b["name"], b["category_he"], b["city"], b["address"],
            b.get("phone") or "לאימות", wa,
            b["website_status"], b["source_name"], b["source_url"],
            f"sites/{b['slug']}/index.html",
        ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(c in (5, 8)))
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3F4F6")
        # קישורים לחיצים
        ws.cell(row=r, column=10).hyperlink = b["source_url"]
        ws.cell(row=r, column=10).font = Font(color="2563EB", underline="single")
        if wa:
            ws.cell(row=r, column=7).hyperlink = wa
            ws.cell(row=r, column=7).font = Font(color="2563EB", underline="single")

    widths = [4, 30, 26, 12, 30, 15, 16, 34, 20, 30, 26]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    # גיליון סיכום
    ws2 = wb.create_sheet("סיכום")
    ws2.sheet_view.rightToLeft = True
    ws2.append(["ריכוז לפי עיר", ""])
    ws2["A1"].font = Font(bold=True, size=13)
    from collections import Counter
    by_city = Counter(b["city"] for b in businesses)
    by_cat = Counter(b["category_he"] for b in businesses)
    ws2.append(["עיר", "מספר עסקים"])
    for city, cnt in by_city.most_common():
        ws2.append([city, cnt])
    ws2.append([])
    ws2.append(["תחום", "מספר עסקים"])
    for cat, cnt in by_cat.most_common():
        ws2.append([cat, cnt])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16

    wb.save(path)


# ----------------------------------------------------------------------------
# main
# ----------------------------------------------------------------------------
def main():
    with open(DATA, encoding="utf-8") as f:
        businesses = json.load(f)

    os.makedirs(SITES, exist_ok=True)
    for b in businesses:
        d = os.path.join(SITES, b["slug"])
        os.makedirs(d, exist_ok=True)
        with open(os.path.join(d, "index.html"), "w", encoding="utf-8") as f:
            f.write(render_site(b))

    with open(os.path.join(SITES, "index.html"), "w", encoding="utf-8") as f:
        f.write(render_index(businesses))

    build_excel(businesses, os.path.join(BASE, "leads.xlsx"))

    print(f"✅ נבנו {len(businesses)} אתרים ב-{SITES}")
    print(f"✅ עמוד אינדקס: {os.path.join(SITES, 'index.html')}")
    print(f"✅ קובץ אקסל: {os.path.join(BASE, 'leads.xlsx')}")


if __name__ == "__main__":
    main()
